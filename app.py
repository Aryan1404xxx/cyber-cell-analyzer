import streamlit as st
import pandas as pd
import numpy as np
from rapidfuzz import fuzz
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import re

st.set_page_config(page_title="Cyber Cell Subject Analyzer", page_icon="🔍", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background-color: #F7F8FA; }
.main-header {
    background: linear-gradient(135deg, #1a3a5c, #2563a8);
    padding: 2rem 2.5rem; border-radius: 16px; margin-bottom: 2rem;
}
.main-header h1 { color: white; font-size: 2rem; font-weight: 700; margin: 0; }
.main-header p  { color: rgba(255,255,255,0.75); margin: 0.3rem 0 0; font-size: 0.95rem; }
.metric-card {
    background: white; border: 1px solid #E5E7EB; border-radius: 12px;
    padding: 1.2rem 1.5rem; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.metric-num   { font-size: 2rem; font-weight: 700; color: #1a3a5c; }
.metric-label { font-size: 0.8rem; color: #6B7280; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
.section-header {
    font-size: 1rem; font-weight: 600; color: #1a3a5c;
    border-left: 4px solid #2563a8; padding-left: 10px; margin: 1.5rem 0 1rem;
}
.badge-exact   { background:#FEE2E2; color:#DC2626; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:600; }
.badge-similar { background:#FEF3C7; color:#D97706; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:600; }
.badge-unique  { background:#D1FAE5; color:#059669; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:600; }
.badge-court   { background:#DBEAFE; color:#1D4ED8; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:600; }
.badge-other   { background:#F3F4F6; color:#374151; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:600; }
.result-card {
    background: white; border: 1px solid #E5E7EB; border-radius: 10px;
    padding: 1rem 1.25rem; margin-bottom: 0.75rem; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
section[data-testid="stSidebar"] { background: #1a3a5c; }
section[data-testid="stSidebar"] * { color: white !important; }
.stButton > button {
    background: #2563a8; color: white; border: none;
    border-radius: 8px; font-weight: 600; padding: 0.5rem 1.5rem;
}
.stButton > button:hover { background: #1a3a5c; color: white; }
.stTabs [aria-selected="true"] { color: #2563a8 !important; border-bottom: 3px solid #2563a8 !important; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>🔍 Cyber Cell Subject Analyzer</h1>
    <p>Upload LEA case data — AI classifies court order vs other cases, then detects duplicates</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("## ⚙️ Settings")
    similarity_threshold = st.slider("Similarity threshold (%)", 50, 95, 65)
    ai_confidence = st.slider("AI classification confidence (%)", 50, 95, 75,
        help="Cases below this confidence get sent to AI for classification")
    st.markdown("---")
    st.markdown("**Classification Method**")
    st.markdown("1️⃣ Keywords scan first")
    st.markdown("2️⃣ AI decides uncertain ones")
    st.markdown("---")
    st.markdown("**How it works**")
    st.markdown("🔴 **Exact** — identical subject")
    st.markdown("🟡 **Similar** — same topic, diff wording")
    st.markdown("🟢 **Unique** — no duplicates")
    st.markdown("🔵 **Court Order** — court related case")
    st.markdown("⚪ **Other** — non-court case")
    st.markdown("---")
    st.markdown("**Developed by:** Aryan Sinha")

COURT_KEYWORDS = [
    "court", "order", "compliance", "fir", "bnss", "ipc", "bns",
    "judge", "judicial", "magistrate", "tribunal", "arbitration", "drt", "drat",
    "summon", "notice", "frozen", "freeze", "lien", "hold", "release",
    "complainant", "petitioner", "respondent", "appeal", "hearing",
    "ps special cell", "special cell", "cyber cell", "police station",
    "e-fir", "efir", "case no", "case number", "u/s", "section",
    "ndoh", "next date", "appearance", "comply", "immediate release",
    "refund", "victim", "alleged", "cheated", "amount"
]

def classify_by_keywords(subject):
    subject_lower = str(subject).lower().strip()
    if not subject_lower or subject_lower in ['nan', 'none', '']:
        return "Other", 1.0
    score = sum(1 for kw in COURT_KEYWORDS if kw in subject_lower)
    total = len(COURT_KEYWORDS)
    confidence = score / total
    if confidence >= 0.05:
        return "Court Order", confidence
    return "Other", 1 - confidence

def classify_by_ai(subjects_uncertain):
    try:
        from sentence_transformers import SentenceTransformer
        from sklearn.metrics.pairwise import cosine_similarity

        court_examples = [
            "Court order compliance FIR 18/26",
            "Release of frozen funds as per court order",
            "Notice U/S 94 BNSS to appear before court",
            "Compliance with Hon'ble court direction",
            "Summon to appear for hearing PS Special Cell",
        ]
        other_examples = [
            "KYC document request for account holder",
            "Account information required for investigation",
            "Transaction details of suspicious account",
            "Request for CCTV footage",
            "Mobile number details required",
        ]

        model = SentenceTransformer('all-MiniLM-L6-v2')
        court_emb  = model.encode(court_examples)
        other_emb  = model.encode(other_examples)
        subj_emb   = model.encode(subjects_uncertain)

        court_center = court_emb.mean(axis=0, keepdims=True)
        other_center = other_emb.mean(axis=0, keepdims=True)

        results = {}
        for i, subj in enumerate(subjects_uncertain):
            emb = subj_emb[i:i+1]
            sim_court = cosine_similarity(emb, court_center)[0][0]
            sim_other = cosine_similarity(emb, other_center)[0][0]
            if sim_court >= sim_other:
                results[subj] = ("Court Order", float(sim_court))
            else:
                results[subj] = ("Other", float(sim_other))
        return results, None
    except Exception as e:
        return {s: ("Other", 0.5) for s in subjects_uncertain}, str(e)

@st.cache_data
def process_file(file_bytes, threshold, ai_conf):
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = [c.strip() for c in df.columns]

    if 'Case ID' not in df.columns or 'Subject' not in df.columns:
        return None, None, None, None, None, None, "Missing required columns: Case ID, Subject"

    df['Case ID'] = df['Case ID'].astype(str).str.strip()
    df['Subject'] = df['Subject'].astype(str).str.strip()

    # Step 1: keyword classification
    classifications = []
    uncertain = []
    for subj in df['Subject']:
        label, conf = classify_by_keywords(subj)
        classifications.append((label, conf))
        if conf < (ai_conf / 100):
            uncertain.append(subj)

    # Step 2: AI for uncertain ones
    ai_results = {}
    ai_error = None
    if uncertain:
        ai_results, ai_error = classify_by_ai(list(set(uncertain)))

    final_labels = []
    final_confs  = []
    for subj, (kw_label, kw_conf) in zip(df['Subject'], classifications):
        if kw_conf < (ai_conf / 100) and subj in ai_results:
            label, conf = ai_results[subj]
        else:
            label, conf = kw_label, kw_conf
        final_labels.append(label)
        final_confs.append(round(conf * 100, 1))

    df['Category']   = final_labels
    df['Confidence'] = final_confs

    court_df = df[df['Category'] == 'Court Order'].copy().reset_index(drop=True)
    other_df = df[df['Category'] == 'Other'].copy().reset_index(drop=True)

    # Duplicate detection
    def get_duplicates(data):
        if data.empty:
            return pd.DataFrame(), pd.DataFrame(), []
        exact_mask = data.duplicated(subset='Subject', keep=False)
        exact = data[exact_mask].copy().sort_values('Subject').reset_index(drop=True)
        grp_map = {}
        for subj, grp in exact.groupby('Subject'):
            ids = ', '.join(grp['Case ID'].tolist())
            for i in grp.index:
                grp_map[i] = ids
        if not exact.empty:
            exact['Matching Case IDs'] = exact.index.map(grp_map)

        unique_data = data.drop_duplicates(subset='Subject').reset_index(drop=True)
        subjects = unique_data['Subject'].tolist()
        fuzzy_groups = []
        visited = set()
        for i in range(len(subjects)):
            if i in visited:
                continue
            group = [i]
            for j in range(i+1, len(subjects)):
                if j in visited:
                    continue
                if fuzz.token_sort_ratio(subjects[i], subjects[j]) >= threshold:
                    group.append(j)
                    visited.add(j)
            if len(group) > 1:
                fuzzy_groups.append(group)
            visited.add(i)

        fuzzy_rows = []
        for g_num, group in enumerate(fuzzy_groups, 1):
            for idx in group:
                subj = subjects[idx]
                matches = data[data['Subject'] == subj]
                for _, row in matches.iterrows():
                    fuzzy_rows.append({
                        'Group': g_num,
                        'Case ID': row['Case ID'],
                        'LEA Name': row.get('LEA Name', ''),
                        'Subject': row['Subject'],
                    })
        fuzzy = pd.DataFrame(fuzzy_rows) if fuzzy_rows else pd.DataFrame()
        return exact, fuzzy, fuzzy_groups

    court_exact, court_fuzzy, court_fgroups = get_duplicates(court_df)
    other_exact, other_fuzzy, other_fgroups = get_duplicates(other_df)

    return df, court_df, other_df, \
           (court_exact, court_fuzzy, court_fgroups), \
           (other_exact, other_fuzzy, other_fgroups), \
           ai_error, None

def build_excel(df, court_df, other_df, court_dupe_data, other_dupe_data):
    court_exact, court_fuzzy, court_fgroups = court_dupe_data
    other_exact, other_fuzzy, other_fgroups = other_dupe_data

    wb = openpyxl.Workbook()

    BLUE_H  = PatternFill("solid", fgColor="1a3a5c")
    COURT_H = PatternFill("solid", fgColor="1D4ED8")
    OTHER_H = PatternFill("solid", fgColor="374151")
    RED_H   = PatternFill("solid", fgColor="DC2626")
    ORG_H   = PatternFill("solid", fgColor="D97706")
    ALT1    = PatternFill("solid", fgColor="EFF6FF")
    ALT2    = PatternFill("solid", fgColor="FFFBEB")
    ALT3    = PatternFill("solid", fgColor="FEF2F2")
    WHITE   = PatternFill("solid", fgColor="FFFFFF")
    GRAY    = PatternFill("solid", fgColor="F9FAFB")

    hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    dfont  = Font(name='Arial', size=10)
    tfont  = Font(name='Arial', bold=True, color='FFFFFF', size=13)
    border = Border(*[Side(style='thin', color='E5E7EB')]*4)

    def hrow(ws, r, fill, n):
        for c in range(1, n+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill; cell.font = hfont
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def drow(ws, r, fill, n):
        for c in range(1, n+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill; cell.font = dfont
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border

    def write_sheet(ws, title, data, header_fill, alt_fill, cols):
        total_cols = len(cols)
        ws.merge_cells(f'A1:{chr(64+total_cols)}1')
        ws['A1'] = title
        ws['A1'].fill = header_fill; ws['A1'].font = tfont
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28
        for c, h in enumerate(cols, 1):
            ws.cell(row=2, column=c, value=h)
        hrow(ws, 2, header_fill, total_cols)
        tog = True
        for i, (_, row) in enumerate(data.iterrows(), 3):
            tog = not tog if i % 2 == 0 else tog
            fill = alt_fill if tog else WHITE
            for c, col in enumerate(cols, 1):
                ws.cell(row=i, column=c, value=str(row.get(col, '')))
            drow(ws, i, fill, total_cols)
            ws.row_dimensions[i].height = 30
        ws.freeze_panes = 'A3'

    # Sheet 1: Court Orders
    ws1 = wb.active; ws1.title = "Court Order Cases"
    cols1 = [c for c in ['Case ID','LEA Name','LEA Email','Subject','Mail Date','Status','Confidence'] if c in court_df.columns or c == 'Confidence']
    write_sheet(ws1, f'COURT ORDER CASES — {len(court_df)} cases', court_df, COURT_H, ALT1, cols1)
    for col, w in zip(['A','B','C','D','E','F','G'], [12,26,30,60,14,12,12]):
        ws1.column_dimensions[col].width = w

    # Sheet 2: Other Cases
    ws2 = wb.create_sheet("Other Cases")
    cols2 = [c for c in ['Case ID','LEA Name','LEA Email','Subject','Mail Date','Status','Confidence'] if c in other_df.columns or c == 'Confidence']
    write_sheet(ws2, f'OTHER CASES — {len(other_df)} cases', other_df, OTHER_H, GRAY, cols2)
    for col, w in zip(['A','B','C','D','E','F','G'], [12,26,30,60,14,12,12]):
        ws2.column_dimensions[col].width = w

    # Sheet 3: Court Order Duplicates
    if not court_exact.empty:
        ws3 = wb.create_sheet("Court Order Duplicates")
        cols3 = ['Case ID','LEA Name','Subject','Matching Case IDs']
        write_sheet(ws3, f'COURT ORDER EXACT DUPLICATES — {len(court_exact)} cases', court_exact, RED_H, ALT3, cols3)
        for col, w in zip(['A','B','C','D'], [12,26,60,28]):
            ws3.column_dimensions[col].width = w

    # Sheet 4: Other Duplicates
    if not other_exact.empty:
        ws4 = wb.create_sheet("Other Duplicates")
        cols4 = ['Case ID','LEA Name','Subject','Matching Case IDs']
        write_sheet(ws4, f'OTHER CASES EXACT DUPLICATES — {len(other_exact)} cases', other_exact, ORG_H, ALT2, cols4)
        for col, w in zip(['A','B','C','D'], [12,26,60,28]):
            ws4.column_dimensions[col].width = w

    # Sheet 5: Summary
    ws5 = wb.create_sheet("Summary")
    ws5.merge_cells('A1:B1')
    ws5['A1'] = 'ANALYSIS SUMMARY'
    ws5['A1'].fill = BLUE_H; ws5['A1'].font = tfont
    ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 28

    rows = [
        ('Total Cases Analyzed', len(df)),
        ('Court Order Cases', len(court_df)),
        ('Other Cases', len(other_df)),
        ('Court Order — Exact Duplicates', len(court_exact)),
        ('Court Order — Similar Groups', len(court_fgroups)),
        ('Other — Exact Duplicates', len(other_exact)),
        ('Other — Similar Groups', len(other_fgroups)),
    ]
    for i, (label, val) in enumerate(rows, 2):
        ws5.cell(row=i, column=1, value=label).font = Font(name='Arial', bold=True, size=11)
        ws5.cell(row=i, column=2, value=val).font = Font(name='Arial', size=11)
        fill = GRAY if i % 2 == 0 else WHITE
        for c in [1, 2]:
            ws5.cell(row=i, column=c).fill = fill
            ws5.cell(row=i, column=c).border = border
            ws5.cell(row=i, column=c).alignment = Alignment(vertical='center')
        ws5.row_dimensions[i].height = 22
    ws5.column_dimensions['A'].width = 40
    ws5.column_dimensions['B'].width = 16

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ---- MAIN UI ----
uploaded = st.file_uploader("📂 Upload Excel file (.xlsx)", type=["xlsx"])

if uploaded is None:
    st.info("👆 Upload an Excel file to get started")
    st.markdown("### 📋 Required Columns")
    sample = pd.DataFrame({
        'Case ID': ['31626', '194560', '287803'],
        'LEA Name': ['Cyber Crime Cell', 'CBI', 'Cyber Crime Cell'],
        'LEA Email': ['acp@delhi.gov.in', 'acp@delhi.gov.in', 'acp@delhi.gov.in'],
        'Subject': ['Court order compliance FIR 18/26', 'Release of frozen funds per court', 'KYC documents required'],
    })
    st.dataframe(sample, use_container_width=True)

else:
    file_bytes = uploaded.read()
    with st.spinner("🤖 Classifying cases and detecting duplicates..."):
        result = process_file(file_bytes, similarity_threshold, ai_confidence)
        df, court_df, other_df, court_dupe_data, other_dupe_data, ai_error, error = result

    if error:
        st.error(f"❌ {error}")
    else:
        if ai_error:
            st.warning(f"⚠️ AI model unavailable ({ai_error}) — using keyword classification only")

        court_exact, court_fuzzy, court_fgroups = court_dupe_data
        other_exact, other_fuzzy, other_fgroups = other_dupe_data

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="metric-card"><div class="metric-num">{len(df)}</div><div class="metric-label">Total Cases</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#1D4ED8">{len(court_df)}</div><div class="metric-label">Court Order Cases</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#374151">{len(other_df)}</div><div class="metric-label">Other Cases</div></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="metric-card"><div class="metric-num" style="color:#DC2626">{len(court_exact)+len(other_exact)}</div><div class="metric-label">Exact Duplicates</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        tab1, tab2, tab3, tab4 = st.tabs(["🔵 Court Order Cases", "⚪ Other Cases", "🔴 Duplicates", "🔎 Search"])

        with tab1:
            st.markdown(f'<div class="section-header">Court Order Cases — {len(court_df)} cases</div>', unsafe_allow_html=True)
            col_a, col_b = st.columns(2)
            col_a.metric("Exact Duplicates", len(court_exact))
            col_b.metric("Similar Groups", len(court_fgroups))

            if not court_exact.empty:
                st.markdown("#### 🔴 Exact Duplicates")
                for subj, grp in court_exact.groupby('Subject'):
                    with st.expander(f"[{len(grp)} cases] {subj[:80]}{'...' if len(subj)>80 else ''}"):
                        st.markdown(f"**Case IDs:** `{', '.join(grp['Case ID'].tolist())}`")
                        show = [c for c in ['Case ID','LEA Name','Mail Date','Status'] if c in grp.columns]
                        st.dataframe(grp[show].reset_index(drop=True), use_container_width=True)

            if not court_fuzzy.empty:
                st.markdown("#### 🟡 Similar Subject Groups")
                for g_num in court_fuzzy['Group'].unique():
                    grp = court_fuzzy[court_fuzzy['Group'] == g_num]
                    preview = grp['Subject'].iloc[0]
                    with st.expander(f"Group {int(g_num)} — {len(grp)} cases — {preview[:60]}..."):
                        for _, row in grp.iterrows():
                            st.markdown(f"• `{row['Case ID']}` — {row['Subject'][:100]}")

            st.markdown("#### 📋 All Court Order Cases")
            show_cols = [c for c in ['Case ID','LEA Name','Subject','Confidence'] if c in court_df.columns]
            st.dataframe(court_df[show_cols], use_container_width=True, height=300)

        with tab2:
            st.markdown(f'<div class="section-header">Other Cases — {len(other_df)} cases</div>', unsafe_allow_html=True)
            col_a, col_b = st.columns(2)
            col_a.metric("Exact Duplicates", len(other_exact))
            col_b.metric("Similar Groups", len(other_fgroups))

            if not other_exact.empty:
                st.markdown("#### 🔴 Exact Duplicates")
                for subj, grp in other_exact.groupby('Subject'):
                    with st.expander(f"[{len(grp)} cases] {subj[:80]}{'...' if len(subj)>80 else ''}"):
                        st.markdown(f"**Case IDs:** `{', '.join(grp['Case ID'].tolist())}`")
                        show = [c for c in ['Case ID','LEA Name','Mail Date','Status'] if c in grp.columns]
                        st.dataframe(grp[show].reset_index(drop=True), use_container_width=True)

            if not other_fuzzy.empty:
                st.markdown("#### 🟡 Similar Subject Groups")
                for g_num in other_fuzzy['Group'].unique():
                    grp = other_fuzzy[other_fuzzy['Group'] == g_num]
                    preview = grp['Subject'].iloc[0]
                    with st.expander(f"Group {int(g_num)} — {len(grp)} cases — {preview[:60]}..."):
                        for _, row in grp.iterrows():
                            st.markdown(f"• `{row['Case ID']}` — {row['Subject'][:100]}")

            st.markdown("#### 📋 All Other Cases")
            show_cols = [c for c in ['Case ID','LEA Name','Subject','Confidence'] if c in other_df.columns]
            st.dataframe(other_df[show_cols], use_container_width=True, height=300)

        with tab3:
            st.markdown('<div class="section-header">All Duplicate Cases</div>', unsafe_allow_html=True)
            st.markdown("#### 🔵 Court Order Exact Duplicates")
            if court_exact.empty:
                st.success("No exact duplicates in court order cases!")
            else:
                show = [c for c in ['Case ID','LEA Name','Subject','Matching Case IDs'] if c in court_exact.columns]
                st.dataframe(court_exact[show], use_container_width=True)

            st.markdown("#### ⚪ Other Cases Exact Duplicates")
            if other_exact.empty:
                st.success("No exact duplicates in other cases!")
            else:
                show = [c for c in ['Case ID','LEA Name','Subject','Matching Case IDs'] if c in other_exact.columns]
                st.dataframe(other_exact[show], use_container_width=True)

        with tab4:
            st.markdown('<div class="section-header">Search by Subject or Case ID</div>', unsafe_allow_html=True)
            query = st.text_input("🔎 Type a subject keyword or Case ID", placeholder="e.g. FIR 18/26 or 287803")
            if query:
                q = query.strip().lower()
                results = df[
                    df['Subject'].str.lower().str.contains(q, na=False) |
                    df['Case ID'].str.lower().str.contains(q, na=False)
                ]
                if results.empty:
                    st.warning("No matching cases found.")
                else:
                    st.success(f"Found {len(results)} matching cases")
                    for _, row in results.iterrows():
                        cat_badge = '<span class="badge-court">Court Order</span>' if row['Category'] == 'Court Order' else '<span class="badge-other">Other</span>'
                        all_exact = pd.concat([court_exact, other_exact]) if not court_exact.empty or not other_exact.empty else pd.DataFrame()
                        all_fuzzy = pd.concat([court_fuzzy, other_fuzzy]) if not court_fuzzy.empty or not other_fuzzy.empty else pd.DataFrame()
                        is_exact = row['Case ID'] in all_exact['Case ID'].values if not all_exact.empty else False
                        is_sim   = row['Case ID'] in all_fuzzy['Case ID'].values if not all_fuzzy.empty else False
                        dupe_badge = '<span class="badge-exact">Exact Duplicate</span>' if is_exact else \
                                     '<span class="badge-similar">Similar Match</span>' if is_sim else \
                                     '<span class="badge-unique">Unique</span>'
                        st.markdown(f"""
                        <div class="result-card">
                            <b>Case ID:</b> {row['Case ID']} &nbsp; {cat_badge} &nbsp; {dupe_badge}<br>
                            <b>LEA:</b> {row.get('LEA Name','')} &nbsp;|&nbsp; <b>Confidence:</b> {row.get('Confidence','')}%<br>
                            <b>Subject:</b> {row['Subject']}
                        </div>
                        """, unsafe_allow_html=True)

        st.markdown("---")
        excel_buf = build_excel(df, court_df, other_df, court_dupe_data, other_dupe_data)
        st.download_button(
            "📥 Download Full Report (Excel)",
            excel_buf,
            "cyber_cell_analysis.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
